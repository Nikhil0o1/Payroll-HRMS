"""Excel report generation."""
from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.attendance import AttendanceDaily, AttendanceLog
from app.models.employee import Employee
from app.models.enums import EmployeeStatus, LeaveStatus
from app.models.leave import LeaveRequest, LeaveType
from app.models.payroll import PayrollDetail, PayrollRun


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0F172A")


def _style_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def _autosize(ws) -> None:
    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, 40)


def _book_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- Attendance report (month) ----
def attendance_report(db: Session, *, year: int, month: int) -> bytes:
    last = monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    _style_header(
        ws,
        ["Employee Code", "Name", "Department", "Date", "Status", "First In", "Last Out", "Worked Hours", "Late?", "Missing?"],
    )

    rows = list(
        db.execute(
            select(AttendanceDaily, Employee)
            .join(Employee, Employee.id == AttendanceDaily.employee_id)
            .where(AttendanceDaily.work_date >= start, AttendanceDaily.work_date <= end)
            .order_by(Employee.employee_code, AttendanceDaily.work_date)
        )
    )
    for daily, emp in rows:
        ws.append(
            [
                emp.employee_code,
                emp.full_name,
                emp.department or "",
                daily.work_date.isoformat(),
                daily.status.value,
                daily.first_in.strftime("%Y-%m-%d %H:%M") if daily.first_in else "",
                daily.last_out.strftime("%Y-%m-%d %H:%M") if daily.last_out else "",
                round(daily.worked_minutes / 60, 2),
                "Yes" if daily.is_late else "",
                "Yes" if daily.has_missing_punch else "",
            ]
        )
    _autosize(ws)
    return _book_to_bytes(wb)


# ---- Leave report ----
def leave_report(db: Session, *, year: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leaves"
    _style_header(
        ws,
        ["Employee Code", "Name", "Leave Type", "Start", "End", "Days", "Status", "Reason", "Decision Note"],
    )
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    rows = list(
        db.execute(
            select(LeaveRequest, Employee, LeaveType)
            .join(Employee, Employee.id == LeaveRequest.employee_id)
            .join(LeaveType, LeaveType.id == LeaveRequest.leave_type_id)
            .where(LeaveRequest.start_date >= start, LeaveRequest.start_date <= end)
            .order_by(Employee.employee_code, LeaveRequest.start_date)
        )
    )
    for req, emp, lt in rows:
        ws.append(
            [
                emp.employee_code,
                emp.full_name,
                lt.code,
                req.start_date.isoformat(),
                req.end_date.isoformat(),
                float(req.days),
                req.status.value,
                req.reason or "",
                req.decision_note or "",
            ]
        )
    _autosize(ws)
    return _book_to_bytes(wb)


# ---- Payroll report (single run) ----
def payroll_report(db: Session, *, run_id: int) -> bytes:
    run = db.get(PayrollRun, run_id)
    if not run:
        raise ValueError("Payroll run not found")
    rows = list(
        db.execute(
            select(PayrollDetail, Employee)
            .join(Employee, Employee.id == PayrollDetail.employee_id)
            .where(PayrollDetail.run_id == run_id)
            .order_by(Employee.employee_code)
        )
    )

    # Collect every component code so we can produce one column per component.
    earning_codes: list[str] = []
    deduction_codes: list[str] = []
    for detail, _emp in rows:
        for e in detail.earnings or []:
            if e["code"] not in earning_codes:
                earning_codes.append(e["code"])
        for d in detail.deductions or []:
            if d["code"] not in deduction_codes:
                deduction_codes.append(d["code"])

    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll-{run.period_year}-{run.period_month:02d}"
    headers = (
        ["Employee Code", "Name", "Department", "Working Days", "Payable Days", "LOP Days"]
        + [f"E:{c}" for c in earning_codes]
        + ["Gross"]
        + [f"D:{c}" for c in deduction_codes]
        + ["Total Deductions", "Net Pay"]
    )
    _style_header(ws, headers)
    for detail, emp in rows:
        e_map = {e["code"]: float(e["amount"]) for e in (detail.earnings or [])}
        d_map = {d["code"]: float(d["amount"]) for d in (detail.deductions or [])}
        ws.append(
            [
                emp.employee_code,
                emp.full_name,
                emp.department or "",
                float(detail.working_days),
                float(detail.payable_days),
                float(detail.lop_days),
            ]
            + [round(e_map.get(c, 0), 2) for c in earning_codes]
            + [round(float(detail.gross), 2)]
            + [round(d_map.get(c, 0), 2) for c in deduction_codes]
            + [round(float(detail.total_deductions), 2), round(float(detail.net_pay), 2)]
        )
    _autosize(ws)
    return _book_to_bytes(wb)


# ---- Employees ----
def employees_report(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    _style_header(
        ws,
        [
            "Code",
            "First Name",
            "Last Name",
            "Work Email",
            "Department",
            "Designation",
            "Employment Type",
            "Status",
            "Date of Joining",
            "Date of Exit",
        ],
    )
    for emp in db.scalars(select(Employee).order_by(Employee.employee_code)):
        ws.append(
            [
                emp.employee_code,
                emp.first_name,
                emp.last_name,
                emp.work_email,
                emp.department or "",
                emp.designation or "",
                emp.employment_type.value,
                emp.status.value,
                emp.date_of_joining.isoformat(),
                emp.date_of_exit.isoformat() if emp.date_of_exit else "",
            ]
        )
    _autosize(ws)
    return _book_to_bytes(wb)
